import tkinter as tk
import openpyxl
from PIL import ImageTk, Image
from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter



def afficher_resultat():
    texte_saisi = str(texte_entry.get())
    nb_pieces=int(texte_entry2.get())
    wb = load_workbook('données chiffrées boites.xlsx')
    ws = wb['inventaire']
    wa = wb['juillet']
    A = get_column_letter(1)#référence
    B = get_column_letter(2)#NAL
    C = get_column_letter(3)#boîte
    D = get_column_letter(4)# nb pièces prises
    boite="non recensée"
    for i in range(1, 10):
        if ws[C + str(i)].value == texte_saisi :
            boite = texte_saisi
            if wa[D + str(i)].value==None :
                wa[D + str(i)].value=0
            cellule_ancienne_valeur = int(wa[D + str(i)].value)
            cellule_nouvelle_valeur = nb_pieces
            somme = cellule_ancienne_valeur + cellule_nouvelle_valeur
            wa[D + str(i)].value = somme
            wb.save('données chiffrées boites.xlsx')
            break


fenetre = tk.Tk()
fenetre.title("programme mise à jour stock")
fenetre.attributes("-fullscreen", True)

# Chargement de l'image
image = Image.open("Logo_Dassault_Aviation.png")  # Remplacez "image.png" par le chemin de votre image
image = image.resize((180, 80))  # Redimensionnez l'image selon vos besoins
image = ImageTk.PhotoImage(image)

# Création d'un label pour l'image
image_label = tk.Label(fenetre, image=image)
image_label.pack(anchor=tk.NE, padx=10, pady=10)

texte_label = tk.Label(fenetre, font=("Arial", 20),text="numéro de la boîte :")
texte_label.pack(pady=10)

texte_entry = tk.Entry(fenetre, font=("Arial", 50), width=5)
texte_entry.pack(pady=10)

texte_label2 = tk.Label(fenetre, font=("Arial", 20),text="nombre de pièces :")
texte_label2.pack(pady=20)

texte_entry2 = tk.Entry(fenetre, font=("Arial", 20), width=5)
texte_entry2.pack(pady=20)

bouton_valider = tk.Button(fenetre,font=("Arial", 20),text="Valider", command=afficher_resultat)
bouton_valider.pack(padx=20, pady=20)

resultat_label = tk.Label(fenetre, font=("Arial", 40), text="")
resultat_label.pack(pady=150)

fenetre.mainloop()



